"""
KAP Bildirim Detay Parse Scripti
=================================
Tek bir disclosure_index uzerinden bildirim detayini ceker ve parse eder.
"""

import json
import sys
import os
import httpx
import re

# Windows encoding fix
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

DISCLOSURE_INDEX = 1556913

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json, text/html, */*",
    "Accept-Language": "tr-TR,tr;q=0.9,en;q=0.8",
    "Referer": "https://www.kap.org.tr/",
}


def try_json_api(client, index):
    """KAP JSON API endpoint denemesi."""
    
    endpoints = [
        f"https://www.kap.org.tr/tr/api/disclosure/{index}",
        f"https://www.kap.org.tr/tr/api/disclosureDetail/{index}",
        f"https://www.kap.org.tr/en/api/disclosure/{index}",
        f"https://www.kap.org.tr/en/api/disclosureDetail/{index}",
        f"https://www.kap.org.tr/tr/api/sgbf-notification/{index}",
        f"https://www.kap.org.tr/en/api/sgbf-notification/{index}",
        f"https://www.kap.org.tr/tr/api/notification/{index}",
        f"https://www.kap.org.tr/en/api/notification/{index}",
    ]
    
    results = []
    
    for url in endpoints:
        try:
            resp = client.get(url, headers=HEADERS, timeout=10, follow_redirects=True)
            status = resp.status_code
            content_type = resp.headers.get("content-type", "")
            body = resp.text[:500] if resp.text else ""
            
            is_json = "json" in content_type or body.strip().startswith(("{", "["))
            
            result = {
                "url": url,
                "status": status,
                "content_type": content_type,
                "is_json": is_json,
                "body_length": len(resp.text) if resp.text else 0,
            }
            
            if status == 200 and is_json:
                try:
                    result["data"] = resp.json()
                    result["success"] = True
                except:
                    result["body_preview"] = body
                    result["success"] = False
            else:
                result["body_preview"] = body[:200]
                result["success"] = False
            
            results.append(result)
            
        except Exception as e:
            results.append({"url": url, "error": str(e), "success": False})
    
    return results


def try_excel_export(client, index):
    """Excel export API."""
    url = f"https://www.kap.org.tr/en/api/notification/export/excel/{index}"
    try:
        resp = client.get(url, headers=HEADERS, timeout=15, follow_redirects=True)
        result = {
            "url": url,
            "status": resp.status_code,
            "content_type": resp.headers.get("content-type", ""),
            "size": len(resp.content),
        }
        
        if resp.status_code == 200:
            excel_path = os.path.join("scripts", f"kap_disclosure_{index}.xlsx")
            with open(excel_path, "wb") as f:
                f.write(resp.content)
            result["saved_to"] = excel_path
            
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path)
                sheets_data = {}
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=False):
                        row_data = {}
                        for cell in row:
                            if cell.value is not None:
                                row_data[cell.coordinate] = str(cell.value)
                        if row_data:
                            rows.append(row_data)
                    sheets_data[sheet_name] = {
                        "max_row": ws.max_row,
                        "max_col": ws.max_column,
                        "rows": rows
                    }
                result["sheets"] = sheets_data
                result["success"] = True
            except ImportError:
                result["note"] = "openpyxl not installed"
                result["success"] = False
            except Exception as e:
                result["parse_error"] = str(e)
                result["success"] = False
        else:
            result["success"] = False
            
        return result
    except Exception as e:
        return {"url": url, "error": str(e), "success": False}


def main():
    index = int(sys.argv[1]) if len(sys.argv) > 1 else DISCLOSURE_INDEX
    
    report = {
        "disclosure_index": index,
        "url": f"https://www.kap.org.tr/en/Bildirim/{index}",
    }
    
    with httpx.Client() as client:
        # 1. JSON API dene
        report["json_api_results"] = try_json_api(client, index)
        
        # 2. Excel export
        report["excel_result"] = try_excel_export(client, index)
    
    # Sonuclari dosyaya yaz
    output_path = os.path.join("scripts", f"kap_disclosure_{index}_report.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"Rapor kaydedildi: {output_path}")
    
    # Ozet
    successful_apis = [r for r in report["json_api_results"] if r.get("success")]
    print(f"\nJSON API: {len(successful_apis)} basarili / {len(report['json_api_results'])} toplam")
    
    if successful_apis:
        print(f"Calisan endpoint: {successful_apis[0]['url']}")
    
    excel = report["excel_result"]
    if excel.get("success"):
        print(f"Excel API: BASARILI ({excel.get('size', 0)} bytes)")
        sheets = excel.get("sheets", {})
        for name, info in sheets.items():
            print(f"  Sheet '{name}': {info['max_row']} satir x {info['max_col']} sutun")
    else:
        print(f"Excel API: {excel.get('status', 'HATA')} - {excel.get('error', '')}")


if __name__ == "__main__":
    main()
